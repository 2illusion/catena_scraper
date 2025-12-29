"""
Data export module - Export scraped data to CSV, Excel, JSON
"""

import csv
import json
import logging
from pathlib import Path
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.models import ArticleInfo
from config.settings import OUTPUT_DIR, CSV_OUTPUT, EXCEL_OUTPUT, JSON_OUTPUT

logger = logging.getLogger(__name__)


class DataExporter:
    """Export scraped article data to various formats"""
    
    def __init__(self, articles: List[ArticleInfo], output_dir: str = OUTPUT_DIR):
        self.articles = articles
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_csv(self, filename: str = CSV_OUTPUT) -> str:
        """Export to CSV file"""
        filepath = self.output_dir / filename
        
        if not self.articles:
            logger.warning("No articles to export")
            return str(filepath)
        
        # Define CSV columns
        fieldnames = [
            'title', 'url', 'doi', 'article_id', 'volume', 'issue', 'year',
            'authors', 'received_date', 'revised_date', 'accepted_date',
            'available_online_date', 'version_of_record_date', 'review_days',
            'scraped_at', 'scrape_status', 'error_message'
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for article in self.articles:
                writer.writerow(article.to_csv_row())
        
        logger.info(f"Exported {len(self.articles)} articles to {filepath}")
        return str(filepath)
    
    def export_json(self, filename: str = JSON_OUTPUT) -> str:
        """Export to JSON file"""
        filepath = self.output_dir / filename
        
        data = {
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "total_articles": len(self.articles),
                "source": "CATENA Journal (ScienceDirect)"
            },
            "articles": [a.to_dict() for a in self.articles]
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Exported {len(self.articles)} articles to {filepath}")
        return str(filepath)
    
    def export_excel(self, filename: str = EXCEL_OUTPUT) -> str:
        """Export to Excel file with formatting and statistics"""
        filepath = self.output_dir / filename
        
        wb = Workbook()
        
        # === Sheet 1: All Articles ===
        ws = wb.active
        ws.title = "Articles"
        
        # Headers
        headers = [
            'Title', 'Volume', 'Year', 'Authors', 'Received Date', 
            'Accepted Date', 'Review Days', 'DOI', 'URL', 'Status'
        ]
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, article in enumerate(self.articles, 2):
            ws.cell(row=row_idx, column=1, value=article.title[:100] if article.title else '')
            ws.cell(row=row_idx, column=2, value=article.volume)
            ws.cell(row=row_idx, column=3, value=article.year)
            ws.cell(row=row_idx, column=4, value='; '.join(article.authors) if article.authors else '')
            ws.cell(row=row_idx, column=5, value=article.received_date)
            ws.cell(row=row_idx, column=6, value=article.accepted_date)
            ws.cell(row=row_idx, column=7, value=article.review_days)
            ws.cell(row=row_idx, column=8, value=article.doi)
            ws.cell(row=row_idx, column=9, value=article.url)
            ws.cell(row=row_idx, column=10, value=article.scrape_status)
            
            # Apply border to all cells
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        # Adjust column widths
        column_widths = [60, 10, 8, 40, 18, 18, 12, 30, 50, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # === Sheet 2: Statistics ===
        ws_stats = wb.create_sheet("Statistics")
        
        successful = [a for a in self.articles if a.scrape_status == "success" and a.review_days is not None]
        
        stats_data = [
            ("Metric", "Value"),
            ("Total Articles Scraped", len(self.articles)),
            ("Successful Scrapes", len(successful)),
            ("Failed Scrapes", len(self.articles) - len(successful)),
        ]
        
        if successful:
            review_days = [a.review_days for a in successful]
            stats_data.extend([
                ("", ""),
                ("Review Time Statistics", "Days"),
                ("Average Review Time", f"{sum(review_days) / len(review_days):.1f}"),
                ("Minimum Review Time", min(review_days)),
                ("Maximum Review Time", max(review_days)),
                ("Median Review Time", sorted(review_days)[len(review_days) // 2]),
            ])
        
        for row_idx, (metric, value) in enumerate(stats_data, 1):
            ws_stats.cell(row=row_idx, column=1, value=metric)
            ws_stats.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws_stats.cell(row=row_idx, column=1).font = header_font
                ws_stats.cell(row=row_idx, column=1).fill = header_fill
                ws_stats.cell(row=row_idx, column=2).font = header_font
                ws_stats.cell(row=row_idx, column=2).fill = header_fill
        
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15
        
        # === Sheet 3: Review Time by Year ===
        ws_by_year = wb.create_sheet("By Year")
        
        # Group by year
        year_stats = {}
        for article in successful:
            year = article.year
            if year not in year_stats:
                year_stats[year] = []
            year_stats[year].append(article.review_days)
        
        ws_by_year.cell(row=1, column=1, value="Year").font = header_font
        ws_by_year.cell(row=1, column=1).fill = header_fill
        ws_by_year.cell(row=1, column=2, value="Article Count").font = header_font
        ws_by_year.cell(row=1, column=2).fill = header_fill
        ws_by_year.cell(row=1, column=3, value="Avg Review Days").font = header_font
        ws_by_year.cell(row=1, column=3).fill = header_fill
        ws_by_year.cell(row=1, column=4, value="Min Days").font = header_font
        ws_by_year.cell(row=1, column=4).fill = header_fill
        ws_by_year.cell(row=1, column=5, value="Max Days").font = header_font
        ws_by_year.cell(row=1, column=5).fill = header_fill
        
        for row_idx, (year, days_list) in enumerate(sorted(year_stats.items()), 2):
            ws_by_year.cell(row=row_idx, column=1, value=year)
            ws_by_year.cell(row=row_idx, column=2, value=len(days_list))
            ws_by_year.cell(row=row_idx, column=3, value=round(sum(days_list) / len(days_list), 1))
            ws_by_year.cell(row=row_idx, column=4, value=min(days_list))
            ws_by_year.cell(row=row_idx, column=5, value=max(days_list))
        
        # Add chart if we have enough data
        if len(year_stats) >= 2:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Average Review Time by Year"
            chart.y_axis.title = "Days"
            chart.x_axis.title = "Year"
            
            data = Reference(ws_by_year, min_col=3, min_row=1, max_row=len(year_stats) + 1)
            cats = Reference(ws_by_year, min_col=1, min_row=2, max_row=len(year_stats) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws_by_year.add_chart(chart, "G2")
        
        # === Sheet 4: Review Time Distribution ===
        ws_dist = wb.create_sheet("Distribution")
        
        if successful:
            # Create distribution buckets (0-30, 31-60, 61-90, etc.)
            buckets = {}
            for article in successful:
                days = article.review_days
                bucket = (days // 30) * 30
                bucket_label = f"{bucket}-{bucket + 29}"
                if bucket_label not in buckets:
                    buckets[bucket_label] = 0
                buckets[bucket_label] += 1
            
            ws_dist.cell(row=1, column=1, value="Review Days Range").font = header_font
            ws_dist.cell(row=1, column=1).fill = header_fill
            ws_dist.cell(row=1, column=2, value="Article Count").font = header_font
            ws_dist.cell(row=1, column=2).fill = header_fill
            
            # Sort buckets by the starting number
            sorted_buckets = sorted(buckets.items(), key=lambda x: int(x[0].split('-')[0]))
            
            for row_idx, (bucket_label, count) in enumerate(sorted_buckets, 2):
                ws_dist.cell(row=row_idx, column=1, value=bucket_label)
                ws_dist.cell(row=row_idx, column=2, value=count)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported {len(self.articles)} articles to {filepath}")
        return str(filepath)
    
    def export_all(self) -> dict:
        """Export to all formats"""
        return {
            'csv': self.export_csv(),
            'json': self.export_json(),
            'excel': self.export_excel()
        }
