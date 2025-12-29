# -*- coding: utf-8 -*-
"""
CATENA Journal Scraper - Fixed Detection Version
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from dateutil import parser as date_parser

import time
import random
import re
import csv
import logging
import argparse
import os
import glob
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def find_chromedriver():
    """Find ChromeDriver in common locations"""
    possible_paths = [
        # Current directory
        "./chromedriver.exe",
        "./chromedriver-win64/chromedriver.exe",
        # webdriver-manager cache
        os.path.expanduser("~/.wdm/drivers/chromedriver/win64/*/chromedriver-win32/chromedriver.exe"),
        os.path.expanduser("~/.wdm/drivers/chromedriver/win64/*/chromedriver.exe"),
    ]
    
    for pattern in possible_paths:
        matches = glob.glob(pattern)
        if matches:
            matches.sort(key=os.path.getmtime, reverse=True)
            return matches[0]
    return None


@dataclass
class ArticleInfo:
    title: str
    url: str
    volume: int = None
    year: int = None
    doi: str = None
    received_date: str = None
    revised_date: str = None
    accepted_date: str = None
    available_online_date: str = None
    review_days: int = None
    scrape_status: str = "pending"
    error_message: str = None
    
    def calculate_review_days(self):
        if self.received_date and self.accepted_date:
            try:
                received = date_parser.parse(self.received_date)
                accepted = date_parser.parse(self.accepted_date)
                self.review_days = (accepted - received).days
            except Exception as e:
                logger.warning(f"Could not parse dates: {e}")


class CatenaScraper:
    BASE_URL = "https://www.sciencedirect.com"
    JOURNAL_URL = "https://www.sciencedirect.com/journal/catena/issues"
    
    def __init__(self):
        self.driver = None
        self.articles: List[ArticleInfo] = []
        Path("data").mkdir(exist_ok=True)
        Path("data/debug").mkdir(parents=True, exist_ok=True)
    
    def init_driver(self):
        """Initialize Chrome WebDriver"""
        driver_path = find_chromedriver()
        
        if not driver_path or not os.path.exists(driver_path):
            logger.error("ChromeDriver not found!")
            logger.info("Please put chromedriver.exe in current directory")
            raise FileNotFoundError("ChromeDriver not found")
        
        logger.info(f"Using ChromeDriver: {driver_path}")
        
        options = Options()
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        
        service = Service(executable_path=driver_path)
        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.implicitly_wait(15)
        
        self.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined});'
        })
        
        logger.info("Chrome initialized!")
    
    def close_driver(self):
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None
            logger.info("Driver closed")
    
    def human_delay(self, min_sec: float = 3, max_sec: float = 8):
        time.sleep(random.uniform(min_sec, max_sec))
    
    def human_scroll(self):
        for _ in range(random.randint(2, 4)):
            scroll = random.randint(100, 400)
            self.driver.execute_script(f"window.scrollBy(0, {scroll});")
            time.sleep(random.uniform(0.3, 0.8))
    
    def is_blocked(self) -> bool:
        """
        Check if page is blocked - STRICT detection only for real block pages.
        Only returns True if we're certain it's a block page.
        """
        try:
            src = self.driver.page_source
            page_title = self.driver.title.lower() if self.driver.title else ""
            
            # Method 1: Check for PerimeterX captcha element (most reliable)
            if 'px-captcha' in src:
                logger.debug("Detected: px-captcha")
                return True
            
            # Method 2: Check for specific Chinese block page text
            # "There was a problem providing the content you requested"
            chinese_block = '\u63d0\u4f9b\u60a8\u8bf7\u6c42\u7684\u5185\u5bb9\u65f6\u51fa\u73b0\u95ee\u9898'
            if chinese_block in src:
                logger.debug("Detected: Chinese block text")
                return True
            
            # Method 3: Check for "Are you a robot?" text (exact phrase)
            if 'are you a robot?' in src.lower():
                logger.debug("Detected: Are you a robot?")
                return True
            
            # Method 4: Check page title for block indicators
            if 'robot' in page_title or 'blocked' in page_title or 'denied' in page_title:
                logger.debug(f"Detected: Block title - {page_title}")
                return True
            
            # Method 5: Very short page with reference number (Chinese block page pattern)
            # Chinese "Reference number"
            chinese_ref = '\u53c2\u8003\u7f16\u53f7'
            if len(src) < 5000 and chinese_ref in src:
                logger.debug("Detected: Short page with reference number")
                return True
            
            return False
            
        except Exception as e:
            logger.debug(f"is_blocked check failed: {e}")
            return False
    
    def wait_if_blocked(self, context: str = "page"):
        """Wait for manual resolution if blocked"""
        if self.is_blocked():
            logger.warning("=" * 60)
            logger.warning(f"BLOCKED on {context}!")
            logger.warning("")
            logger.warning("In the browser window:")
            logger.warning("1. Wait 30-60 seconds for auto-redirect")
            logger.warning("2. Or click/hold the verification button")
            logger.warning("3. Or refresh the page (F5)")
            logger.warning("")
            logger.warning("Press Enter when you see normal content...")
            logger.warning("=" * 60)
            input()
            time.sleep(2)
            
            if self.is_blocked():
                logger.warning("Still blocked. Enter=retry, skip=skip, quit=stop")
                response = input().strip().lower()
                if response == 'skip':
                    return False
                elif response == 'quit':
                    raise KeyboardInterrupt("User quit")
                else:
                    return self.wait_if_blocked(context)
        return True
    
    def save_debug_page(self, filename: str):
        filepath = Path('data/debug') / filename
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(self.driver.page_source)
        logger.info(f"Saved debug: {filepath}")
    
    def get_volumes(self, max_volumes: int = 1) -> List[dict]:
        logger.info("Opening journal issues page...")
        self.driver.get(self.JOURNAL_URL)
        self.human_delay(3, 5)
        
        if not self.wait_if_blocked("journal page"):
            return []
        
        self.human_scroll()
        
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        volumes = []
        
        for link in soup.find_all('a', href=re.compile(r'/journal/catena/vol/\d+')):
            href = link.get('href', '')
            match = re.search(r'/vol/(\d+)', href)
            if not match:
                continue
            
            vol_num = int(match.group(1))
            url = self.BASE_URL + href if href.startswith('/') else href
            
            if not any(v['volume'] == vol_num for v in volumes):
                volumes.append({'volume': vol_num, 'year': 2025, 'url': url})
        
        volumes.sort(key=lambda x: x['volume'], reverse=True)
        logger.info(f"Found {len(volumes)} volumes")
        return volumes[:max_volumes]
    
    def get_articles(self, volume: dict, max_articles: int = 10) -> List[ArticleInfo]:
        logger.info(f"Opening Volume {volume['volume']}...")
        self.driver.get(volume['url'])
        self.human_delay(3, 5)
        
        if not self.wait_if_blocked("volume page"):
            return []
        
        self.human_scroll()
        
        # Scroll to load all articles
        for _ in range(3):
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
        
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        articles = []
        seen = set()
        
        for link in soup.find_all('a', href=re.compile(r'/science/article/pii/')):
            href = link.get('href', '')
            url = self.BASE_URL + href if href.startswith('/') else href
            
            if url in seen:
                continue
            seen.add(url)
            
            title = link.get_text(strip=True)
            if len(title) < 10:
                continue
            
            articles.append(ArticleInfo(
                title=title[:200],
                url=url,
                volume=volume['volume'],
                year=volume['year']
            ))
        
        logger.info(f"Found {len(articles)} articles")
        return articles[:max_articles]
    
    def extract_dates(self, soup: BeautifulSoup, text: str, article: ArticleInfo):
        """Extract dates from article page"""
        
        # Pattern: "Received 24 April 2025, Revised 10 November 2025, Accepted 20 November 2025"
        date_patterns = [
            (r'Received\s+(\d{1,2}\s+\w+\s+\d{4})', 'received_date'),
            (r'Revised\s+(\d{1,2}\s+\w+\s+\d{4})', 'revised_date'),
            (r'Accepted\s+(\d{1,2}\s+\w+\s+\d{4})', 'accepted_date'),
            (r'Available\s+online\s+(\d{1,2}\s+\w+\s+\d{4})', 'available_online_date'),
        ]
        
        for pattern, field in date_patterns:
            if not getattr(article, field):
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    setattr(article, field, match.group(1))
                    logger.info(f"  {field}: {match.group(1)}")
        
        # Get DOI
        doi_meta = soup.find('meta', {'name': 'citation_doi'})
        if doi_meta:
            article.doi = doi_meta.get('content')
    
    def get_article_details(self, article: ArticleInfo, num: int) -> ArticleInfo:
        logger.info(f"[{num}] {article.title[:50]}...")
        
        try:
            self.driver.get(article.url)
            self.human_delay(2, 4)
            
            if not self.wait_if_blocked(f"article {num}"):
                article.scrape_status = "skipped"
                return article
            
            self.human_scroll()
            
            # Scroll to load content
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            self.driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            
            # Save first article for debugging
            if num == 1:
                self.save_debug_page("first_article.html")
            
            # Check for block after content load
            if self.is_blocked():
                if not self.wait_if_blocked("article content"):
                    article.scrape_status = "blocked"
                    return article
            
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            text = soup.get_text()
            
            self.extract_dates(soup, text, article)
            article.calculate_review_days()
            
            if article.received_date or article.accepted_date:
                article.scrape_status = "success"
                if article.review_days:
                    logger.info(f"  Review time: {article.review_days} days")
            else:
                article.scrape_status = "no_dates"
                logger.warning(f"  No dates found")
                self.save_debug_page(f"no_dates_{num}.html")
            
            return article
            
        except KeyboardInterrupt:
            raise
        except Exception as e:
            logger.error(f"Error: {e}")
            article.scrape_status = "error"
            article.error_message = str(e)
            return article
    
    def export_results(self):
        if not self.articles:
            logger.warning("No articles to export")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # CSV
        csv_file = Path('data') / f'catena_{timestamp}.csv'
        with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
            fields = ['title', 'volume', 'year', 'received_date', 'accepted_date',
                     'review_days', 'doi', 'url', 'scrape_status']
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            for a in self.articles:
                writer.writerow({k: getattr(a, k) for k in fields})
        logger.info(f"Saved: {csv_file}")
        
        # Excel
        xlsx_file = Path('data') / f'catena_{timestamp}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = "Articles"
        
        headers = ['Title', 'Volume', 'Year', 'Received', 'Accepted', 
                   'Review Days', 'DOI', 'URL', 'Status']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        for a in self.articles:
            ws.append([a.title, a.volume, a.year, a.received_date, a.accepted_date,
                      a.review_days, a.doi, a.url, a.scrape_status])
        
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['H'].width = 50
        wb.save(xlsx_file)
        logger.info(f"Saved: {xlsx_file}")
        
        # Stats
        total = len(self.articles)
        success = sum(1 for a in self.articles if a.scrape_status == "success")
        with_dates = sum(1 for a in self.articles if a.review_days)
        
        logger.info(f"\n{'='*60}")
        logger.info(f"RESULTS: {total} total, {success} success, {with_dates} with dates")
        if with_dates:
            avg = sum(a.review_days for a in self.articles if a.review_days) / with_dates
            logger.info(f"Average review time: {avg:.1f} days")
            logger.info("=" * 60)
    
    def scrape(self, max_volumes: int = 1, max_articles: int = 10):
        try:
            self.init_driver()
            
            # Visit homepage first
            logger.info("Visiting homepage...")
            self.driver.get(self.BASE_URL)
            self.human_delay(3, 5)
            
            if not self.wait_if_blocked("homepage"):
                logger.error("Cannot access homepage")
                return
            
            self.human_scroll()
            
            # Get volumes
            volumes = self.get_volumes(max_volumes)
            if not volumes:
                logger.error("No volumes found")
                return
            
            # Process volumes
            for vol in volumes:
                logger.info(f"\n{'='*60}")
                logger.info(f"Volume {vol['volume']}")
                logger.info('='*60)
                
                articles = self.get_articles(vol, max_articles)
                
                for i, article in enumerate(articles, 1):
                    try:
                        article = self.get_article_details(article, i)
                        self.articles.append(article)
                        self.human_delay(5, 10)
                    except KeyboardInterrupt:
                        break
            
            self.export_results()
            
        except KeyboardInterrupt:
            logger.info("Interrupted, saving results...")
            self.export_results()
        finally:
            self.close_driver()


def main():
    parser = argparse.ArgumentParser(description='CATENA Scraper')
    parser.add_argument('--test', action='store_true', help='Test mode (1 vol, 5 articles)')
    parser.add_argument('--max-volumes', type=int, default=1)
    parser.add_argument('--max-articles', type=int, default=10)
    
    args = parser.parse_args()
    
    if args.test:
        args.max_volumes = 1
        args.max_articles = 5
    
    print("\n" + "="*60)
    print("CATENA Journal Scraper")
    print("="*60 + "\n")
    
    scraper = CatenaScraper()
    scraper.scrape(args.max_volumes, args.max_articles)


if __name__ == '__main__':
    main()